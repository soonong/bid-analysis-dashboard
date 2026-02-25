import openpyxl
import json
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\imbid\Desktop\안티그래비티용\발주공고분석\랄랄라.xlsx')
ws = wb.active

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print('Headers:', headers)
print('Total rows:', ws.max_row - 1)

data = []
for r in range(2, ws.max_row + 1):
    row = {}
    for c, h in enumerate(headers, 1):
        val = ws.cell(r, c).value
        if isinstance(val, datetime):
            val = val.strftime('%Y-%m-%d')
        row[h] = val
    data.append(row)

print('\nAll data:')
for d in data:
    print(d)

# Save as JSON
with open(r'C:\Users\imbid\Desktop\안티그래비티용\발주공고분석\data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print('\nSaved to data.json')
