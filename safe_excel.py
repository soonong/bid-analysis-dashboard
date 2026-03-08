import openpyxl
import pandas as pd
import json
import traceback

try:
    print("Loading workbook...")
    wb = openpyxl.load_workbook('12월 전기.xlsx', data_only=True, read_only=True)
    ws = wb.active
    
    print("Finding headers...")
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            headers.append(str(cell.value) if cell.value else '')
            
    bid_idx = -1
    for i, h in enumerate(headers):
        if '공고번호' in h or '번호' in h:
            bid_idx = i
            break
            
    print(f"Header found at index: {bid_idx}")
    
    if bid_idx == -1:
        print("Could not find bid number column!")
        print("Headers:", headers)
    else:
        user_bids = set()
        print("Reading rows...")
        for count, row in enumerate(ws.iter_rows(min_row=2)):
            val = row[bid_idx].value
            if val:
                user_bids.add(str(val).strip())
            
            if count > 5000:
                print("Too many rows, breaking early!")
                break
                
        print(f"Loaded {len(user_bids)} user bids.")
        
        print("Loading CSV...")
        my_df = pd.read_csv('2025_12_Seoul_Electric.csv')
        my_bids = set(my_df['공고번호'].astype(str).str.strip())
        
        print(f"User bids: {len(user_bids)}, My bids: {len(my_bids)}")
        only_user = user_bids - my_bids
        print("Missing in My CSV (Only in User Excel):", len(only_user))
        
        if only_user:
            print("Loading data.json to find missing items...")
            with open('data.json', 'r', encoding='utf-8') as f:
                data = json.load(f)
            found = 0
            for r in data:
                bid = r.get('공고번호', '').strip()
                if bid in only_user:
                    print(f"Missing Item -> {bid} | Date:{r.get('입력일')} | Reg:{r.get('지역제한')} | Cat:{r.get('종목')} | Amt:{r.get('추정가격',0):,}")
                    found += 1
            print(f"Found {found} missing items in data.json.")
except Exception as e:
    print("Error:", e)
    traceback.print_exc()
